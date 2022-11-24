import os

from openpyxl import load_workbook
import telebot
from random import randint


keyboard = telebot.types.ReplyKeyboardMarkup(True)
key_get_quote = telebot.types.InlineKeyboardButton (text='Дай цитату!')
keyboard.add(key_get_quote)
bot = telebot.TeleBot("943087428:AAEoLmSLXJfsTHbkA-wRIEmhoGKb-8SPrxI")
database_file = os.path.join(os.getcwd(), 'db.xlsx')


def get_quote():
	_load = load_workbook(filename=database_file).active
	max_row = _load.max_row

	random_int = randint(1, max_row)
	quote = _load.cell(row=random_int, column=1).value
	return quote


def write_quote(text=None):
	wb = load_workbook(database_file)
	sheet = wb.active

	_load = load_workbook(filename=database_file).active
	row_to_write = _load.max_row + 1

	cell_to_write = sheet.cell(row=row_to_write, column=1)
	cell_to_write.value = text

	wb.save(filename=database_file)
	wb.close()


@bot.message_handler(func=lambda message: True)
def talk(message):
	if message.text == 'Дай цитату!':
		bot.reply_to(message, get_quote())
	elif message.text == 'Привет!':
		bot.send_message(message.chat.id, "Привет!", reply_markup=keyboard)
	else:
		write_quote(message.text)
		bot.reply_to(message, "Записал!")


bot.polling()

# if __name__ == '__main__':
#
# 	print(getQuote())
# 	writeQuote("uuuФФФФФ")




